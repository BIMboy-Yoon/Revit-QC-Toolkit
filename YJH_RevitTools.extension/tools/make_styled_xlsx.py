# -*- coding: utf-8 -*-

from __future__ import print_function

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_NAVY = "34495A"
TITLE_NAVY = "3C4A57"
TEXT_NAVY = "263645"
SOFT_NAVY_LINE = "405060"
ORANGE_POINT = "E97826"
LIGHT_BORDER = "D9DEE3"
LIGHT_FILL = "F6F7F8"
ZEBRA_FILL = "F2F4F6"


def load_payload(json_path):
    with open(json_path, "r", encoding="utf-8") as input_file:
        return json.load(input_file)


def has_suit_font():
    try:
        import winreg

        font_key_path = r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Fonts"
        font_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, font_key_path)
        value_index = 0

        while True:
            try:
                value_name, value_data, value_type = winreg.EnumValue(
                    font_key,
                    value_index
                )
                combined_text = "{0} {1}".format(value_name, value_data).lower()
                if "suit" in combined_text:
                    winreg.CloseKey(font_key)
                    return True
                value_index += 1
            except OSError:
                break

        winreg.CloseKey(font_key)
    except Exception:
        pass

    return False


def get_report_font_name():
    if has_suit_font():
        return "SUIT"
    return "Malgun Gothic"


def build_styles():
    font_name = get_report_font_name()
    thin_side = Side(style="thin", color=LIGHT_BORDER)

    return {
        "title_font": Font(
            name=font_name,
            size=16,
            bold=True,
            color="FFFFFF"
        ),
        "header_font": Font(
            name=font_name,
            size=10,
            bold=True,
            color="FFFFFF"
        ),
        "body_font": Font(name=font_name, size=9, color=TEXT_NAVY),
        "body_bold_font": Font(
            name=font_name,
            size=9,
            bold=True,
            color=TEXT_NAVY
        ),
        "summary_label_font": Font(
            name=font_name,
            size=9,
            bold=True,
            color=TEXT_NAVY
        ),
        "highlight_font": Font(
            name=font_name,
            size=12,
            bold=True,
            color=TEXT_NAVY
        ),
        "title_fill": PatternFill(fill_type="solid", fgColor=TITLE_NAVY),
        "navy_fill": PatternFill(fill_type="solid", fgColor=HEADER_NAVY),
        "orange_fill": PatternFill(fill_type="solid", fgColor=ORANGE_POINT),
        "zebra_fill": PatternFill(fill_type="solid", fgColor=ZEBRA_FILL),
        "summary_label_fill": PatternFill(
            fill_type="solid",
            fgColor=LIGHT_FILL
        ),
        "white_fill": PatternFill(fill_type="solid", fgColor="FFFFFF"),
        "highlight_fill": PatternFill(fill_type="solid", fgColor="FFF0E3"),
        "high_fill": PatternFill(fill_type="solid", fgColor="FADBD8"),
        "medium_fill": PatternFill(fill_type="solid", fgColor="FFF0C2"),
        "low_fill": PatternFill(fill_type="solid", fgColor="E9ECEF"),
        "thin_border": Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        ),
        "title_alignment": Alignment(
            horizontal="left",
            vertical="center"
        ),
        "header_alignment": Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        ),
        "body_alignment": Alignment(vertical="top", wrap_text=True)
    }


def apply_title(sheet, title, column_count, styles):
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=column_count
    )
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = styles["title_alignment"]
    sheet.row_dimensions[1].height = 32

    for column_index in range(1, column_count + 1):
        accent_cell = sheet.cell(row=2, column=column_index)
        accent_cell.fill = styles["orange_fill"]
    sheet.row_dimensions[2].height = 4


def auto_fit_columns(sheet, column_count, last_row):
    for column_index in range(1, column_count + 1):
        max_length = 0

        for row_index in range(1, last_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            value_length = len(str(value))
            if value_length > max_length:
                max_length = value_length

        adjusted_width = max(12, min(max_length + 3, 52))
        sheet.column_dimensions[get_column_letter(column_index)].width = (
            adjusted_width
        )


def apply_table(
    sheet,
    title,
    headers,
    rows,
    styles,
    severity_column,
    count_column=None
):
    column_count = len(headers)
    apply_title(sheet, title, column_count, styles)
    header_row = 3
    data_start_row = 4

    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["navy_fill"]
        cell.border = styles["thin_border"]
        cell.alignment = styles["header_alignment"]

    for row_offset, row_values in enumerate(rows):
        row_index = data_start_row + row_offset
        zebra_fill = styles["zebra_fill"] if row_offset % 2 == 1 else None

        for column_index, value in enumerate(row_values, 1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = styles["body_font"]
            cell.border = styles["thin_border"]
            cell.alignment = styles["body_alignment"]

            if zebra_fill is not None:
                cell.fill = zebra_fill
            if count_column is not None and column_index == count_column:
                cell.font = styles["body_bold_font"]

        severity_value = str(row_values[severity_column - 1])
        severity_cell = sheet.cell(row=row_index, column=severity_column)
        severity_cell.font = styles["body_bold_font"]

        if severity_value == "High":
            severity_cell.fill = styles["high_fill"]
        elif severity_value == "Medium":
            severity_cell.fill = styles["medium_fill"]
        elif severity_value == "Low":
            severity_cell.fill = styles["low_fill"]

    last_row = max(header_row, data_start_row + len(rows) - 1)
    last_column = get_column_letter(column_count)
    sheet.auto_filter.ref = "A{0}:{1}{2}".format(
        header_row,
        last_column,
        last_row
    )
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    auto_fit_columns(sheet, column_count, last_row)


def write_summary_sheet(sheet, payload, styles):
    summary_data = payload.get("summary_data", {})
    metadata = payload.get("metadata", {})
    apply_title(sheet, "Revit QC Report", 4, styles)
    sheet.sheet_view.showGridLines = False

    summary_items = [
        ("Project", metadata.get("project", "")),
        ("Active Config", metadata.get("active_config", "")),
        ("Run Mode", metadata.get("run_mode", "")),
        ("QC Status", metadata.get("qc_status", "")),
        ("Checked Sheets", summary_data.get("checked_sheets", 0)),
        ("Checked Views", summary_data.get("checked_views", 0)),
        (
            "Checked Parameter Elements",
            metadata.get("checked_parameter_elements", 0)
        ),
        ("Total Review Items", summary_data.get("total_issues", 0)),
        ("Review Groups", metadata.get("review_group_count", 0)),
        ("High", summary_data.get("high_count", 0)),
        ("Medium", summary_data.get("medium_count", 0)),
        ("Low", summary_data.get("low_count", 0)),
        ("Run Time", metadata.get("run_time", "")),
        ("Export Time", metadata.get("export_time", "")),
        ("Tool Version", metadata.get("tool_version", "")),
        ("Export Folder", metadata.get("export_folder", ""))
    ]

    for item_index, item in enumerate(summary_items):
        card_row = 4 + int(item_index / 2)
        card_column = 1 + (item_index % 2) * 2
        label_cell = sheet.cell(row=card_row, column=card_column, value=item[0])
        value_cell = sheet.cell(
            row=card_row,
            column=card_column + 1,
            value=item[1]
        )

        label_cell.font = styles["summary_label_font"]
        label_cell.fill = styles["summary_label_fill"]
        label_cell.border = styles["thin_border"]
        label_cell.alignment = styles["body_alignment"]
        value_cell.font = styles["body_bold_font"]
        value_cell.fill = styles["white_fill"]
        value_cell.border = styles["thin_border"]
        value_cell.alignment = styles["body_alignment"]

        if item[0] in ["Total Review Items", "Review Groups"]:
            value_cell.fill = styles["highlight_fill"]
            value_cell.font = styles["highlight_font"]

    sheet.column_dimensions["A"].width = 27
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 27
    sheet.column_dimensions["D"].width = 42

    for row_index in range(4, 12):
        sheet.row_dimensions[row_index].height = 34
    sheet.row_dimensions[11].height = 42
    sheet.freeze_panes = "A4"


def create_workbook(payload):
    styles = build_styles()
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "QC Summary"
    write_summary_sheet(summary_sheet, payload, styles)

    review_groups_sheet = workbook.create_sheet("Review Groups")
    apply_table(
        review_groups_sheet,
        "Review Groups",
        [
            "Category",
            "Item Type",
            "QC Item",
            "Severity",
            "Count",
            "Sample Items",
            "Recommendation"
        ],
        payload.get("review_groups", []),
        styles,
        4,
        5
    )

    key_samples_sheet = workbook.create_sheet("Key Samples")
    apply_table(
        key_samples_sheet,
        "Key Samples",
        [
            "Category",
            "Item Type",
            "Item Name",
            "Severity",
            "QC Item",
            "Review Message",
            "Recommendation"
        ],
        payload.get("key_samples", []),
        styles,
        4
    )

    full_detail_sheet = workbook.create_sheet("Full Detail")
    apply_table(
        full_detail_sheet,
        "Full Detail",
        [
            "Category",
            "Severity",
            "ElementId",
            "ElementType",
            "Name",
            "QC Item",
            "Review Message",
            "Current Value",
            "Recommendation"
        ],
        payload.get("full_detail", []),
        styles,
        2
    )

    return workbook


def main(arguments):
    if len(arguments) != 3:
        print(
            "Usage: make_styled_xlsx.py <input_json_path> <output_xlsx_path>",
            file=sys.stderr
        )
        return 2

    input_json_path = os.path.abspath(arguments[1])
    output_xlsx_path = os.path.abspath(arguments[2])

    try:
        payload = load_payload(input_json_path)
        workbook = create_workbook(payload)
        workbook.save(output_xlsx_path)
        print(output_xlsx_path)
        return 0
    except Exception as ex:
        print("Styled XLSX helper failed: {0}".format(ex), file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
